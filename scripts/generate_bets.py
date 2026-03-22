"""
Ladder Bet Intelligence Engine
Uses Claude extended thinking (Opus) to analyse every fixture
and output a ranked, reasoned bet feed — JSON + Excel.
"""

import os
import json
import requests
import anthropic
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle

# ── Config ────────────────────────────────────────────────────────────────────

ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
ODDS_API_KEY      = os.environ.get("ODDS_API_KEY", "")
TODAY             = datetime.now(timezone.utc).strftime("%Y-%m-%d")
OUTPUT_JSON       = f"docs/data/bets.json"
OUTPUT_EXCEL      = f"docs/data/bets_{TODAY}.xlsx"
HISTORY_JSON      = "docs/data/history.json"

# Sports to pull — add/remove as needed
SPORTS = [
    "soccer_epl",
    "soccer_germany_bundesliga",
    "soccer_italy_serie_a",
    "soccer_spain_la_liga",
    "basketball_nba",
]

# Markets to request
MARKETS = "h2h,spreads,totals"
REGIONS = "uk,eu,au"

# ── Odds API fetch ─────────────────────────────────────────────────────────────

def fetch_fixtures(sport: str) -> list[dict]:
    """Fetch upcoming odds from The Odds API. Falls back to empty on error."""
    if not ODDS_API_KEY:
        return []
    url = f"https://api.the-odds-api.com/v4/sports/{sport}/odds"
    params = {
        "apiKey": ODDS_API_KEY,
        "regions": REGIONS,
        "markets": MARKETS,
        "oddsFormat": "decimal",
        "dateFormat": "iso",
    }
    try:
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"  [warn] {sport}: {e}")
        return []


def flatten_fixtures(raw_fixtures: list[dict], sport: str) -> list[dict]:
    """Turn raw Odds API response into a clean list of opportunities."""
    opportunities = []
    for game in raw_fixtures:
        home = game.get("home_team", "")
        away = game.get("away_team", "")
        commence = game.get("commence_time", "")

        # Skip if already started
        try:
            from dateutil.parser import parse as parse_dt
            if parse_dt(commence) < datetime.now(timezone.utc):
                continue
        except Exception:
            pass

        bookmakers = game.get("bookmakers", [])
        if not bookmakers:
            continue

        # Aggregate best available odds across all bookmakers per market
        best_odds: dict[str, dict] = {}

        for bk in bookmakers:
            for market in bk.get("markets", []):
                key = market["key"]
                for outcome in market.get("outcomes", []):
                    name = outcome["name"]
                    price = outcome["price"]
                    entry = best_odds.setdefault(f"{key}::{name}", {
                        "market": key, "outcome": name, "odds": 0
                    })
                    if price > entry["odds"]:
                        entry["odds"] = price

        for opp in best_odds.values():
            if opp["odds"] < 1.15 or opp["odds"] > 12.0:
                continue  # filter noise
            opportunities.append({
                "sport":    sport.replace("soccer_", "").replace("basketball_", "").upper(),
                "home":     home,
                "away":     away,
                "match":    f"{home} vs {away}",
                "commence": commence,
                "market":   opp["market"],
                "outcome":  opp["outcome"],
                "odds":     round(opp["odds"], 2),
            })

    return opportunities


def prefilter_opportunities(opportunities: list[dict]) -> list[dict]:
    """
    Aggressively filter to only ladder-relevant opportunities before
    sending to Claude. Targets odds 1.65-2.30 (our sweet spot for 2x).
    Also deduplicates — max 3 markets per match.
    Goal: 40-60 opportunities max, not 600+.
    """
    # Step 1: Odds range filter — only ladder-relevant prices
    filtered = [o for o in opportunities if 1.65 <= o["odds"] <= 2.30]

    # Step 2: Priority markets — skip low-value markets
    priority_markets = {"h2h", "spreads", "totals", "btts", "double_chance"}
    filtered = [o for o in filtered if any(m in o.get("market", "") for m in priority_markets)]

    # Step 3: Max 3 opportunities per match (best odds per market type)
    from collections import defaultdict
    match_markets: dict[str, list] = defaultdict(list)
    for o in filtered:
        match_markets[o["match"]].append(o)

    deduped = []
    for match, opps in match_markets.items():
        # Sort by how close to 2.0 the odds are (our target)
        opps.sort(key=lambda x: abs(x["odds"] - 2.0))
        deduped.extend(opps[:3])

    # Step 4: Cap total at 60 — sort by proximity to 2.0 odds
    deduped.sort(key=lambda x: abs(x["odds"] - 2.0))
    final = deduped[:60]

    print(f"  [filter] {len(opportunities)} → {len(final)} opportunities after pre-filter")
    return final


# ── Load real stats context ────────────────────────────────────────────────────

def load_football_context() -> dict:
    """Load enriched football fixture data produced by fetch_football_data.py"""
    try:
        with open("docs/data/football_context.json") as f:
            data = json.load(f)
        # Index by match string for quick lookup
        index = {}
        for fix in data.get("fixtures", []):
            key = f"{fix.get('home', '')} vs {fix.get('away', '')}".lower()
            index[key] = fix
        print(f"  [context] Football: {len(index)} enriched fixtures loaded")
        return index
    except (FileNotFoundError, json.JSONDecodeError):
        print("  [context] No football context found — running without stats")
        return {}


def load_nba_context() -> dict:
    """Load enriched NBA game data produced by fetch_nba_data.py"""
    try:
        with open("docs/data/nba_context.json") as f:
            data = json.load(f)
        index = {}
        for game in data.get("games", []):
            key = f"{game.get('home', '')} vs {game.get('away', '')}".lower()
            index[key] = game
        print(f"  [context] NBA: {len(index)} enriched games loaded")
        return index
    except (FileNotFoundError, json.JSONDecodeError):
        print("  [context] No NBA context found — running without stats")
        return {}


def enrich_opportunity(opp: dict, football_ctx: dict, nba_ctx: dict) -> dict:
    """Attach real stats context to an odds opportunity before sending to Claude."""
    match_key = opp.get("match", "").lower()
    sport     = opp.get("sport", "").upper()

    # Try football context first
    ctx = football_ctx.get(match_key)
    if not ctx and "NBA" in sport:
        ctx = nba_ctx.get(match_key)

    # Fuzzy match — try partial home/away name match
    if not ctx:
        home = opp.get("home", "").lower()
        away = opp.get("away", "").lower()
        for key, val in {**football_ctx, **nba_ctx}.items():
            if home[:8] in key and away[:8] in key:
                ctx = val
                break

    if ctx:
        opp["stats_context"] = {
            "home_form":       ctx.get("home_form", {}),
            "away_form":       ctx.get("away_form", {}),
            "h2h":             ctx.get("h2h", {}),
            "home_injuries":   ctx.get("home_injuries", []),
            "away_injuries":   ctx.get("away_injuries", []),
            "home_standing":   ctx.get("home_standing", {}),
            "away_standing":   ctx.get("away_standing", {}),
            # NBA specific
            "home_season_stats": ctx.get("home_season_stats", {}),
            "away_season_stats": ctx.get("away_season_stats", {}),
            "home_splits":     ctx.get("home_splits", {}),
            "away_splits":     ctx.get("away_splits", {}),
            "home_b2b":        ctx.get("home_b2b", False),
            "away_b2b":        ctx.get("away_b2b", False),
        }
    return opp


# ── Claude extended-thinking analysis ─────────────────────────────────────────

SYSTEM_PROMPT = """You are an elite sports betting analyst with access to REAL statistical data.
You are powering a "2x Ladder Challenge" — find the highest-confidence bets targeting ~2.0x returns.

CRITICAL RULES:
- You have been given REAL stats in the "stats_context" field for each opportunity. USE THEM.
- Reference specific numbers in your reasoning — form strings, goals averages, net ratings, injury names
- If stats_context is empty, flag this as increased uncertainty and lower your confidence accordingly
- Do NOT fabricate statistics. If you don't have a number, say so.
- Be ruthlessly honest. A 55% confidence is often more accurate than claiming 75%.

STATS YOU HAVE ACCESS TO (when available):
Football: Last 5 form, goals scored/conceded avg, H2H record, over 2.5 rate, BTTS rate, injuries, standings
NBA: Season net rating, offensive/defensive rating, pace, home/away splits, back-to-back flag, injuries, last 10 form

ANALYSIS FRAMEWORK — work through this for every bet:
1. What does the form data actually say? (not vibes — numbers)
2. What does H2H tell us? (how often has this market landed historically?)
3. Are there injury/suspension concerns that move the needle?
4. Is the bookmaker price fair given the underlying data? (implied prob vs your model prob)
5. What could go wrong? (risk flags must be data-driven, not generic)

For each opportunity output a JSON object with EXACTLY these fields:

{
  "match": "string",
  "sport": "string",
  "commence": "ISO datetime string",
  "market": "string",
  "outcome": "string",
  "odds": float,
  "verdict": "TAKE" | "SKIP" | "MONITOR",
  "confidence": integer (0-100),
  "edge_rating": integer (1-5),
  "expected_value": float (e.g. 0.12 = +12% EV),
  "ladder_fit": "PERFECT" | "GOOD" | "MARGINAL" | "POOR",
  "reasoning": "string — MUST reference specific stats from stats_context. Min 4 sentences.",
  "risk_flags": ["specific", "data-backed", "risk", "factors"],
  "best_combo_partner": "string or null",
  "kelly_fraction": float (0.0-0.25),
  "stats_used": true | false
}

Ladder fit:
- PERFECT: odds 1.90-2.15, confidence >65%, low variance, stats support
- GOOD: odds 1.75-2.30, confidence >55%, stats available
- MARGINAL: borderline confidence OR odds off-target OR missing stats
- POOR: negative EV, high variance, key injuries, no stats

Output ONLY a valid JSON array. No preamble. No explanation outside the JSON."""


def analyse_with_claude(opportunities: list[dict]) -> list[dict]:
    """Send enriched opportunities to Claude Opus with extended thinking."""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    # Build the user message — send in groups of 20 (tight, focused batches)
    chunk_size = 20
    all_results = []

    for i in range(0, len(opportunities), chunk_size):
        chunk = opportunities[i:i + chunk_size]

        user_msg = f"""Today is {TODAY}. Analyse these {len(chunk)} betting opportunities 
and return your full assessment as a JSON array.

OPPORTUNITIES:
{json.dumps(chunk, indent=2)}

Remember: output ONLY the JSON array, nothing else."""

        print(f"  [claude] Analysing batch {i//chunk_size + 1} ({len(chunk)} opportunities)...")

        response = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=16000,
            thinking={
                "type": "adaptive",
            },
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_msg}],
        )

        # Extract text blocks (skip thinking blocks)
        raw_text = ""
        for block in response.content:
            if block.type == "text":
                raw_text = block.text
                break

        try:
            parsed = json.loads(raw_text)
            all_results.extend(parsed)
        except json.JSONDecodeError:
            # Try to salvage with bracket extraction
            start = raw_text.find("[")
            end   = raw_text.rfind("]") + 1
            if start != -1 and end > start:
                try:
                    parsed = json.loads(raw_text[start:end])
                    all_results.extend(parsed)
                except Exception:
                    print(f"  [warn] Could not parse Claude response for batch {i}")

    return all_results


# ── Excel builder ──────────────────────────────────────────────────────────────

# Colour palette
C_DARK_BG    = "1A1A2E"
C_HEADER_BG  = "16213E"
C_ACCENT     = "0F3460"
C_TAKE_BG    = "0D4A0D"
C_TAKE_FG    = "52D452"
C_SKIP_BG    = "4A0D0D"
C_SKIP_FG    = "FF6B6B"
C_MON_BG     = "4A3A0D"
C_MON_FG     = "FFD166"
C_WHITE      = "FFFFFF"
C_LIGHT_GRAY = "B0B8C8"
C_GOLD       = "FFD700"
C_SUBHEADER  = "0A2342"

def _font(bold=False, size=10, color=C_WHITE, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin", color="2A2A4A"):
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def build_excel(analysed: list[dict], date_str: str) -> None:
    wb = Workbook()

    # ── Sheet 1: Today's Best Bets ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "🎯 Best Bets Today"
    ws1.sheet_view.showGridLines = False

    # Sort: TAKE first, then MONITOR, then SKIP; within each by confidence desc
    order = {"TAKE": 0, "MONITOR": 1, "SKIP": 2}
    takes    = [b for b in analysed if b.get("verdict") == "TAKE"]
    monitors = [b for b in analysed if b.get("verdict") == "MONITOR"]
    skips    = [b for b in analysed if b.get("verdict") == "SKIP"]

    for lst in [takes, monitors, skips]:
        lst.sort(key=lambda x: x.get("confidence", 0), reverse=True)

    best_bets = takes[:10] + monitors[:5] + skips[:3]

    # Title row
    ws1.merge_cells("A1:N1")
    ws1["A1"] = f"⚡  LADDER BET INTELLIGENCE  —  {date_str}"
    ws1["A1"].font      = _font(bold=True, size=16, color=C_GOLD)
    ws1["A1"].fill      = _fill(C_DARK_BG)
    ws1["A1"].alignment = _align("center")
    ws1.row_dimensions[1].height = 36

    # Subtitle row
    ws1.merge_cells("A2:N2")
    ws1["A2"] = f"Powered by Claude Opus · Extended Thinking · {len(analysed)} opportunities analysed · {len(takes)} TAKE signals"
    ws1["A2"].font      = _font(size=10, color=C_LIGHT_GRAY, italic=True)
    ws1["A2"].fill      = _fill(C_DARK_BG)
    ws1["A2"].alignment = _align("center")
    ws1.row_dimensions[2].height = 22

    # Column headers
    headers = [
        "Verdict", "Confidence", "Edge", "Ladder Fit",
        "Match", "Sport", "Market", "Outcome",
        "Odds", "EV%", "Kelly %",
        "Reasoning", "Risk Flags", "Combo Partner"
    ]
    col_widths = [10, 13, 8, 13, 36, 12, 14, 22, 8, 8, 9, 60, 35, 28]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.font      = _font(bold=True, size=10, color=C_GOLD)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _align("center")
        cell.border    = _border()
        ws1.column_dimensions[get_column_letter(col_idx)].width = w

    ws1.row_dimensions[3].height = 22
    ws1.freeze_panes = "A4"

    # Data rows
    for row_offset, bet in enumerate(best_bets, 4):
        verdict    = bet.get("verdict", "SKIP")
        confidence = bet.get("confidence", 0)
        edge       = bet.get("edge_rating", 0)
        lfit       = bet.get("ladder_fit", "POOR")

        # Row background alternating
        row_bg = "1E1E3A" if row_offset % 2 == 0 else "16182E"

        # Verdict colours
        if verdict == "TAKE":
            v_bg, v_fg = C_TAKE_BG, C_TAKE_FG
        elif verdict == "MONITOR":
            v_bg, v_fg = C_MON_BG, C_MON_FG
        else:
            v_bg, v_fg = C_SKIP_BG, C_SKIP_FG

        # Ladder fit colours
        lfit_colors = {
            "PERFECT": ("0A3A1F", "52D452"),
            "GOOD":    ("1A2E0A", "8FD44F"),
            "MARGINAL":("3A2A0A", "FFD166"),
            "POOR":    ("3A0A0A", "FF6B6B"),
        }
        lfit_bg, lfit_fg = lfit_colors.get(lfit, ("222222", "AAAAAA"))

        values = [
            verdict,
            f"{confidence}%",
            f"{'★' * edge}{'☆' * (5-edge)}",
            lfit,
            bet.get("match", ""),
            bet.get("sport", ""),
            bet.get("market", "").replace("_", " ").title(),
            bet.get("outcome", ""),
            bet.get("odds", 0),
            f"{round(bet.get('expected_value', 0) * 100, 1)}%",
            f"{round(bet.get('kelly_fraction', 0) * 100, 1)}%",
            bet.get("reasoning", ""),
            ", ".join(bet.get("risk_flags", [])),
            bet.get("best_combo_partner") or "—",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row_offset, column=col_idx, value=val)
            cell.fill   = _fill(row_bg)
            cell.border = _border()
            cell.font   = _font(size=9, color=C_WHITE)
            cell.alignment = _align("left", wrap=True)

            # Verdict cell
            if col_idx == 1:
                cell.fill      = _fill(v_bg)
                cell.font      = _font(bold=True, size=9, color=v_fg)
                cell.alignment = _align("center")

            # Confidence cell — colour gradient
            elif col_idx == 2:
                conf_color = (
                    C_TAKE_FG  if confidence >= 70 else
                    "FFD166"   if confidence >= 55 else
                    C_SKIP_FG
                )
                cell.font      = _font(bold=True, size=9, color=conf_color)
                cell.alignment = _align("center")

            # Ladder fit
            elif col_idx == 4:
                cell.fill      = _fill(lfit_bg)
                cell.font      = _font(bold=True, size=9, color=lfit_fg)
                cell.alignment = _align("center")

            # Odds
            elif col_idx == 9:
                cell.font      = _font(bold=True, size=9, color=C_GOLD)
                cell.alignment = _align("center")
                cell.number_format = "0.00"

        ws1.row_dimensions[row_offset].height = 52

    # ── Sheet 2: Full Analysis ──────────────────────────────────────────────
    ws2 = wb.create_sheet("📊 Full Analysis")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:N1")
    ws2["A1"] = f"FULL OPPORTUNITY ANALYSIS  —  {date_str}  —  {len(analysed)} total opportunities"
    ws2["A1"].font      = _font(bold=True, size=13, color=C_GOLD)
    ws2["A1"].fill      = _fill(C_DARK_BG)
    ws2["A1"].alignment = _align("center")
    ws2.row_dimensions[1].height = 30

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font      = _font(bold=True, size=10, color=C_GOLD)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _align("center")
        cell.border    = _border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    ws2.row_dimensions[2].height = 22
    ws2.freeze_panes = "A3"

    all_sorted = sorted(analysed, key=lambda x: (order.get(x.get("verdict","SKIP"), 2), -x.get("confidence",0)))

    for row_offset, bet in enumerate(all_sorted, 3):
        verdict = bet.get("verdict", "SKIP")
        row_bg  = "1E1E3A" if row_offset % 2 == 0 else "16182E"
        if verdict == "TAKE":
            v_bg, v_fg = C_TAKE_BG, C_TAKE_FG
        elif verdict == "MONITOR":
            v_bg, v_fg = C_MON_BG, C_MON_FG
        else:
            v_bg, v_fg = C_SKIP_BG, C_SKIP_FG

        confidence = bet.get("confidence", 0)
        lfit = bet.get("ladder_fit", "POOR")
        edge = bet.get("edge_rating", 0)
        lfit_bg, lfit_fg = lfit_colors.get(lfit, ("222222","AAAAAA"))

        values = [
            verdict,
            f"{confidence}%",
            f"{'★' * edge}{'☆' * (5-edge)}",
            lfit,
            bet.get("match",""), bet.get("sport",""),
            bet.get("market","").replace("_"," ").title(),
            bet.get("outcome",""),
            bet.get("odds",0),
            f"{round(bet.get('expected_value',0)*100,1)}%",
            f"{round(bet.get('kelly_fraction',0)*100,1)}%",
            bet.get("reasoning",""),
            ", ".join(bet.get("risk_flags",[])),
            bet.get("best_combo_partner") or "—",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_offset, column=col_idx, value=val)
            cell.fill   = _fill(row_bg)
            cell.border = _border()
            cell.font   = _font(size=9, color=C_WHITE)
            cell.alignment = _align("left", wrap=True)
            if col_idx == 1:
                cell.fill      = _fill(v_bg)
                cell.font      = _font(bold=True, size=9, color=v_fg)
                cell.alignment = _align("center")
            elif col_idx == 2:
                conf_color = C_TAKE_FG if confidence >= 70 else "FFD166" if confidence >= 55 else C_SKIP_FG
                cell.font      = _font(bold=True, size=9, color=conf_color)
                cell.alignment = _align("center")
            elif col_idx == 4:
                cell.fill      = _fill(lfit_bg)
                cell.font      = _font(bold=True, size=9, color=lfit_fg)
                cell.alignment = _align("center")
            elif col_idx == 9:
                cell.font      = _font(bold=True, size=9, color=C_GOLD)
                cell.alignment = _align("center")
                cell.number_format = "0.00"
        ws2.row_dimensions[row_offset].height = 48

    # ── Sheet 3: Parlay Builder ─────────────────────────────────────────────
    ws3 = wb.create_sheet("🔗 Parlay Builder")
    ws3.sheet_view.showGridLines = False

    # Top TAKE bets that are "PERFECT" or "GOOD" ladder fit
    parlay_candidates = [b for b in takes if b.get("ladder_fit") in ("PERFECT","GOOD")][:8]

    ws3.merge_cells("A1:H1")
    ws3["A1"] = "PARLAY BUILDER — Combine legs to hit 2.0x target"
    ws3["A1"].font      = _font(bold=True, size=13, color=C_GOLD)
    ws3["A1"].fill      = _fill(C_DARK_BG)
    ws3["A1"].alignment = _align("center")
    ws3.row_dimensions[1].height = 30

    parlay_headers = ["#", "Match", "Outcome", "Odds", "Confidence", "Ladder Fit", "Combo Odds (running)", "Ladder Verdict"]
    parlay_widths  = [5, 36, 22, 8, 13, 13, 22, 16]

    for col_idx, (h, w) in enumerate(zip(parlay_headers, parlay_widths), 1):
        cell = ws3.cell(row=2, column=col_idx, value=h)
        cell.font      = _font(bold=True, size=10, color=C_GOLD)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _align("center")
        cell.border    = _border()
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    running_odds = 1.0
    for row_offset, bet in enumerate(parlay_candidates, 3):
        running_odds *= bet.get("odds", 1.0)
        lfit = bet.get("ladder_fit","POOR")
        lfit_bg, lfit_fg = lfit_colors.get(lfit,("222222","AAAAAA"))
        row_bg = "1E1E3A" if row_offset % 2 == 0 else "16182E"
        ladder_v = "✅ PERFECT" if 1.9 <= running_odds <= 2.15 else "⚡ ABOVE" if running_odds > 2.15 else "⬇ BELOW"

        vals = [
            row_offset - 2,
            bet.get("match",""),
            bet.get("outcome",""),
            bet.get("odds",0),
            f"{bet.get('confidence',0)}%",
            lfit,
            f"{round(running_odds, 3)}x",
            ladder_v,
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_offset, column=col_idx, value=val)
            cell.fill      = _fill(row_bg)
            cell.border    = _border()
            cell.font      = _font(size=9, color=C_WHITE)
            cell.alignment = _align("center")
            if col_idx == 4:
                cell.font = _font(bold=True, size=9, color=C_GOLD)
            if col_idx == 6:
                cell.fill = _fill(lfit_bg)
                cell.font = _font(bold=True, size=9, color=lfit_fg)
            if col_idx == 7:
                cell.font = _font(bold=True, size=10, color=C_TAKE_FG if 1.85 <= running_odds <= 2.20 else C_MON_FG)
        ws3.row_dimensions[row_offset].height = 28

    # ── Sheet 4: Model Log ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("📈 Model Log")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:F1")
    ws4["A1"] = "DAILY MODEL PERFORMANCE LOG"
    ws4["A1"].font      = _font(bold=True, size=13, color=C_GOLD)
    ws4["A1"].fill      = _fill(C_DARK_BG)
    ws4["A1"].alignment = _align("center")
    ws4.row_dimensions[1].height = 30

    log_headers = ["Date", "Total Opportunities", "TAKE Signals", "MONITOR Signals", "Avg Confidence (TAKE)", "Best Edge Bet"]
    log_widths   = [14, 22, 16, 18, 24, 40]
    for col_idx, (h, w) in enumerate(zip(log_headers, log_widths), 1):
        cell = ws4.cell(row=2, column=col_idx, value=h)
        cell.font      = _font(bold=True, size=10, color=C_GOLD)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _align("center")
        cell.border    = _border()
        ws4.column_dimensions[get_column_letter(col_idx)].width = w

    avg_conf = round(sum(b.get("confidence",0) for b in takes) / max(len(takes),1), 1)
    best_edge = max(takes, key=lambda x: x.get("edge_rating",0), default={})

    row_vals = [
        date_str,
        len(analysed),
        len(takes),
        len(monitors),
        f"{avg_conf}%",
        best_edge.get("match","—") + " — " + best_edge.get("outcome",""),
    ]
    for col_idx, val in enumerate(row_vals, 1):
        cell = ws4.cell(row=3, column=col_idx, value=val)
        cell.fill      = _fill("16182E")
        cell.border    = _border()
        cell.font      = _font(size=9, color=C_WHITE)
        cell.alignment = _align("center")

    ws4.row_dimensions[3].height = 24

    wb.save(OUTPUT_EXCEL)
    print(f"  [excel] Saved: {OUTPUT_EXCEL}")


# ── History log ────────────────────────────────────────────────────────────────

def update_history(analysed: list[dict], date_str: str) -> None:
    try:
        with open(HISTORY_JSON) as f:
            history = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        history = []

    takes    = [b for b in analysed if b.get("verdict") == "TAKE"]
    monitors = [b for b in analysed if b.get("verdict") == "MONITOR"]

    history.append({
        "date":          date_str,
        "total":         len(analysed),
        "take_count":    len(takes),
        "monitor_count": len(monitors),
        "top_bets":      takes[:5],
    })

    # Keep last 90 days
    history = history[-90:]

    with open(HISTORY_JSON, "w") as f:
        json.dump(history, f, indent=2)
    print(f"  [history] Updated: {HISTORY_JSON}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*60}")
    print(f"  LADDER BET ENGINE  —  {TODAY}")
    print(f"{'='*60}\n")

    # 1. Fetch odds
    all_opportunities = []
    if ODDS_API_KEY:
        for sport in SPORTS:
            print(f"  [odds] Fetching {sport}...")
            raw = fetch_fixtures(sport)
            opps = flatten_fixtures(raw, sport)
            all_opportunities.extend(opps)
            print(f"         → {len(opps)} opportunities")
    else:
        print("  [warn] No ODDS_API_KEY — using demo data")
        all_opportunities = _demo_opportunities()

    print(f"\n  [total] {len(all_opportunities)} raw opportunities\n")

    if not all_opportunities:
        print("  [exit] No opportunities found.")
        return

    # 2. Pre-filter to ladder-relevant odds before anything else
    all_opportunities = prefilter_opportunities(all_opportunities)

    # 3. Load real stats context
    print("  [context] Loading stats context...")
    football_ctx = load_football_context()
    nba_ctx      = load_nba_context()
    stats_count  = len(football_ctx) + len(nba_ctx)
    print(f"  [context] {stats_count} enriched fixtures available\n")

    # 3. Enrich each opportunity with real stats
    enriched_opps = []
    stats_matched = 0
    for opp in all_opportunities:
        enriched = enrich_opportunity(opp, football_ctx, nba_ctx)
        if enriched.get("stats_context"):
            stats_matched += 1
        enriched_opps.append(enriched)

    print(f"  [enrich] {stats_matched}/{len(enriched_opps)} opportunities matched with real stats\n")

    # 4. Claude analysis with extended thinking
    print("  [model] Starting Claude Opus extended thinking analysis...")
    analysed = analyse_with_claude(enriched_opps)
    print(f"  [model] Analysis complete — {len(analysed)} bets scored\n")

    # 5. Save JSON
    os.makedirs("docs/data", exist_ok=True)
    output = {
        "generated":    TODAY,
        "total":        len(analysed),
        "take_count":   sum(1 for b in analysed if b.get("verdict") == "TAKE"),
        "stats_powered": stats_matched,
        "bets":         sorted(analysed, key=lambda x: (
                            {"TAKE":0,"MONITOR":1,"SKIP":2}.get(x.get("verdict","SKIP"),2),
                            -x.get("confidence",0)
                        )),
    }
    with open(OUTPUT_JSON, "w") as f:
        json.dump(output, f, indent=2)
    print(f"  [json]  Saved: {OUTPUT_JSON}")

    # 6. Build Excel
    build_excel(analysed, TODAY)

    # 7. Update history
    update_history(analysed, TODAY)

    print(f"\n{'='*60}")
    print(f"  Done.")
    print(f"  TAKE signals:   {output['take_count']} / {output['total']}")
    print(f"  Stats-powered:  {stats_matched} / {len(enriched_opps)}")
    print(f"{'='*60}\n")


def _demo_opportunities() -> list[dict]:
    """Demo data when no Odds API key is present — for local testing."""
    return [
        {"sport":"EPL","home":"Arsenal","away":"Chelsea","match":"Arsenal vs Chelsea",
         "commence":"2026-03-22T15:00:00Z","market":"h2h","outcome":"Arsenal","odds":1.85},
        {"sport":"EPL","home":"Arsenal","away":"Chelsea","match":"Arsenal vs Chelsea",
         "commence":"2026-03-22T15:00:00Z","market":"totals","outcome":"Over 2.5","odds":1.72},
        {"sport":"NBA","home":"Boston Celtics","away":"Miami Heat","match":"Boston Celtics vs Miami Heat",
         "commence":"2026-03-22T23:00:00Z","market":"spreads","outcome":"Boston Celtics -5.5","odds":1.91},
        {"sport":"BUNDESLIGA","home":"Bayern Munich","away":"Dortmund","match":"Bayern Munich vs Dortmund",
         "commence":"2026-03-22T17:30:00Z","market":"h2h","outcome":"Bayern Munich","odds":1.60},
        {"sport":"BUNDESLIGA","home":"Bayern Munich","away":"Dortmund","match":"Bayern Munich vs Dortmund",
         "commence":"2026-03-22T17:30:00Z","market":"totals","outcome":"Over 2.5","odds":1.55},
    ]


if __name__ == "__main__":
    main()
